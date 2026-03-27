#!/usr/bin/env python3
"""
data_clean.py — 数据清洗 CLI 工具（Phase 1 MVP）

支持格式: CSV, TSV, JSON, JSONL, Excel (.xlsx)

子命令:
  profile <file>                  数据画像（质量报告）
  execute <file> [--operations]   执行清洗操作
  validate <original> <cleaned>   清洗前后对比验证
  history <file>                  查看清洗版本历史

设计原则:
  - LLM 做判断（通过 profile 输出理解数据），工具做计算
  - 原始数据 immutable，清洗结果写入版本链
  - 每步操作可回滚，审计日志 append-only
  - 输出 JSON 供 LLM 消费，report.md 供人类阅读
"""

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sys
import time
from collections import Counter
from datetime import datetime

# ── 常量 ─────────────────────────────────────────────────

WORKSPACE = os.path.expanduser("~/.data_clean/workspace")
VERSION_DIR = os.path.join(WORKSPACE, "versions")
LOG_FILE = os.path.join(WORKSPACE, "audit.jsonl")

# 日期格式识别模式
DATE_PATTERNS = [
    (r"^\d{4}-\d{2}-\d{2}$", "%Y-%m-%d", "YYYY-MM-DD"),
    (r"^\d{4}/\d{2}/\d{2}$", "%Y/%m/%d", "YYYY/MM/DD"),
    (r"^\d{2}/\d{2}/\d{4}$", "%d/%m/%Y", "DD/MM/YYYY"),
    (r"^\d{1,2}/\d{1,2}/\d{4}$", "%m/%d/%Y", "M/D/YYYY"),
    (r"^\d{1,2}/\d{1,2}/\d{2}$", "%m/%d/%y", "M/D/YY"),
    (r"^[A-Z][a-z]{2}\s+\d{1,2}\s+\d{4}$", "%b %d %Y", "Mon DD YYYY"),
]

PLACEHOLDER_VALUES = {"n/a", "na", "null", "none", "undefined", "tbd", "-", "--", ""}

# 支持的文件格式
SUPPORTED_FORMATS = {"csv", "tsv", "json", "jsonl", "xlsx"}


# ── 工具函数 ─────────────────────────────────────────────

def ensure_workspace():
    """确保工作目录存在"""
    os.makedirs(VERSION_DIR, exist_ok=True)


# ── 格式检测与统一 I/O ──────────────────────────────────

def detect_format(filepath):
    """从文件扩展名推断格式"""
    ext = os.path.splitext(filepath)[1].lower().lstrip(".")
    if ext in ("tsv", "tab"):
        return "tsv"
    if ext == "jsonl":
        return "jsonl"
    if ext == "json":
        # 自动区分 JSON 和 JSONL：JSONL 每行一个 JSON 对象
        try:
            with open(filepath, "r", encoding="utf-8-sig") as f:
                first_char = f.read(1).strip()
                if first_char == "[" or first_char == "{":
                    f.seek(0)
                    lines = [l.strip() for l in f if l.strip()]
                    if len(lines) > 1 and all(_is_json_obj(l) for l in lines[:3]):
                        return "jsonl"
            return "json"
        except Exception:
            return "json"
    if ext in ("xlsx", "xls"):
        return "xlsx"
    return "csv"


def _is_json_obj(line):
    """检查一行是否是独立的 JSON 对象"""
    try:
        obj = json.loads(line)
        return isinstance(obj, dict)
    except (json.JSONDecodeError, ValueError):
        return False


def _stringify(value):
    """将任意值转为字符串（统一内部处理格式）"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def read_data(filepath, fmt=None):
    """统一读取入口，返回 (headers, rows)，rows 为 list[dict[str,str]]"""
    if fmt is None:
        fmt = detect_format(filepath)

    if fmt == "csv":
        return _read_csv(filepath)
    elif fmt == "tsv":
        return _read_tsv(filepath)
    elif fmt == "json":
        return _read_json(filepath)
    elif fmt == "jsonl":
        return _read_jsonl(filepath)
    elif fmt == "xlsx":
        return _read_xlsx(filepath)
    else:
        raise ValueError(f"不支持的格式: {fmt}")


def write_data(filepath, headers, rows, fmt=None):
    """统一写入入口（原子写入）"""
    if fmt is None:
        fmt = detect_format(filepath)

    if fmt == "csv":
        _write_csv(filepath, headers, rows)
    elif fmt == "tsv":
        _write_tsv(filepath, headers, rows)
    elif fmt == "json":
        _write_json(filepath, headers, rows)
    elif fmt == "jsonl":
        _write_jsonl(filepath, headers, rows)
    elif fmt == "xlsx":
        _write_xlsx(filepath, headers, rows)
    else:
        raise ValueError(f"不支持的格式: {fmt}")


# ── CSV ──────────────────────────────────────────────────

def _read_csv(filepath):
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        return [], []
    return list(rows[0].keys()), rows


def _write_csv(filepath, headers, rows):
    tmp = filepath + ".tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    os.replace(tmp, filepath)


# ── TSV ──────────────────────────────────────────────────

def _read_tsv(filepath):
    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter="\t")
        rows = list(reader)
    if not rows:
        return [], []
    return list(rows[0].keys()), rows


def _write_tsv(filepath, headers, rows):
    tmp = filepath + ".tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)
    os.replace(tmp, filepath)


# ── JSON ─────────────────────────────────────────────────

def _read_json(filepath):
    with open(filepath, "r", encoding="utf-8-sig") as f:
        data = json.load(f)

    # 支持: [{...}, {...}] 或 {"data": [{...}, {...}]} 或 {"records": [...]}
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        # 尝试常见的包装键
        for key in ("data", "records", "rows", "items", "results"):
            if key in data and isinstance(data[key], list):
                records = data[key]
                break
        else:
            # 单个对象当作一行
            records = [data]
    else:
        return [], []

    if not records:
        return [], []

    # 提取所有键作为 headers（保持第一条记录的键序）
    headers = list(records[0].keys())
    seen = set(headers)
    for r in records[1:]:
        for k in r.keys():
            if k not in seen:
                headers.append(k)
                seen.add(k)

    # 统一转字符串
    rows = []
    for r in records:
        rows.append({h: _stringify(r.get(h, "")) for h in headers})

    return headers, rows


def _write_json(filepath, headers, rows):
    tmp = filepath + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    os.replace(tmp, filepath)


# ── JSONL ────────────────────────────────────────────────

def _read_jsonl(filepath):
    records = []
    with open(filepath, "r", encoding="utf-8-sig") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError:
                continue

    if not records:
        return [], []

    headers = list(records[0].keys())
    seen = set(headers)
    for r in records[1:]:
        for k in r.keys():
            if k not in seen:
                headers.append(k)
                seen.add(k)

    rows = []
    for r in records:
        rows.append({h: _stringify(r.get(h, "")) for h in headers})

    return headers, rows


def _write_jsonl(filepath, headers, rows):
    tmp = filepath + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    os.replace(tmp, filepath)


# ── Excel (.xlsx) ────────────────────────────────────────

def _read_xlsx(filepath):
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "读取 Excel 需要 openpyxl: pip3 install openpyxl"
        )

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 1:
        wb.close()
        return [], []

    # 读取所有行
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return [], []

    # 寻找表头行：第一个至少有 2 个非空单元格的行
    header_idx = 0
    for i, row in enumerate(all_rows):
        non_empty = sum(1 for v in row if v is not None and str(v).strip())
        if non_empty >= 2:
            header_idx = i
            break

    raw_headers = all_rows[header_idx]
    # 生成列名：有值用原值，空的自动命名
    headers = []
    for i, h in enumerate(raw_headers):
        name = _stringify(h).strip()
        if not name:
            name = f"col_{i+1}"
        headers.append(name)

    # 去重列名（Excel 可能有同名列）
    seen_names = {}
    for i, h in enumerate(headers):
        if h in seen_names:
            seen_names[h] += 1
            headers[i] = f"{h}_{seen_names[h]}"
        else:
            seen_names[h] = 0

    # 读取数据行（跳过表头行及之前的行）
    rows = []
    for raw_row in all_rows[header_idx + 1:]:
        # 跳过全空行
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        row = {}
        for i, h in enumerate(headers):
            val = raw_row[i] if i < len(raw_row) else None
            row[h] = _stringify(val)
        rows.append(row)

    return headers, rows


def _write_xlsx(filepath, headers, rows):
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "写入 Excel 需要 openpyxl: pip3 install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    tmp = filepath + ".tmp.xlsx"
    wb.save(tmp)
    wb.close()
    os.replace(tmp, filepath)


# ── 兼容旧接口（内部使用） ────────────────────────────────

def read_csv(filepath):
    """兼容旧接口：读取 CSV"""
    return _read_csv(filepath)


def write_csv(filepath, headers, rows):
    """兼容旧接口：写入 CSV"""
    _write_csv(filepath, headers, rows)


def file_hash(filepath):
    """计算文件 SHA256"""
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()[:12]


def audit_log(action, details):
    """追加审计日志"""
    ensure_workspace()
    entry = {
        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action,
        **details,
    }
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def save_version(filepath, headers, rows, label):
    """保存数据版本快照（版本链统一用 CSV，体积小且可读）"""
    ensure_workspace()
    basename = os.path.splitext(os.path.basename(filepath))[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    version_file = os.path.join(VERSION_DIR, f"{basename}_{label}_{ts}.csv")
    _write_csv(version_file, headers, rows)
    return version_file


# ── profile 子命令 ───────────────────────────────────────

def infer_date_format(value):
    """推断日期格式"""
    v = value.strip()
    for pattern, fmt, label in DATE_PATTERNS:
        if re.match(pattern, v):
            try:
                datetime.strptime(v, fmt)
                return label
            except ValueError:
                continue
    return None


def detect_column_type(values):
    """推断列的数据类型"""
    non_empty = [v for v in values if v.strip().lower() not in PLACEHOLDER_VALUES]
    if not non_empty:
        return "empty"

    numeric_count = 0
    date_count = 0
    for v in non_empty:
        # 数值检测
        try:
            float(v.replace(",", "").strip())
            numeric_count += 1
            continue
        except ValueError:
            pass
        # 日期检测
        if infer_date_format(v):
            date_count += 1

    total = len(non_empty)
    if numeric_count > total * 0.7:
        return "numeric"
    if date_count > total * 0.5:
        return "datetime"
    return "text"


def profile_column(name, values):
    """生成单列的详细画像"""
    total = len(values)
    non_empty = [v for v in values if v.strip().lower() not in PLACEHOLDER_VALUES and v.strip()]
    missing = total - len(non_empty)

    col_type = detect_column_type(values)
    unique_vals = list(set(non_empty))
    val_counts = Counter(non_empty)

    profile = {
        "name": name,
        "type": col_type,
        "total": total,
        "missing": missing,
        "missing_rate": f"{missing / total * 100:.1f}%" if total > 0 else "0%",
        "unique": len(unique_vals),
        "top_values": val_counts.most_common(8),
    }

    # 类型特定分析
    issues = []

    if col_type == "datetime":
        formats = Counter()
        bad_dates = []
        for v in non_empty:
            fmt = infer_date_format(v)
            if fmt:
                formats[fmt] += 1
            else:
                bad_dates.append(v)
        if len(formats) > 1:
            issues.append({
                "type": "mixed_date_formats",
                "severity": "high",
                "detail": f"检测到 {len(formats)} 种日期格式: {dict(formats)}",
            })
        if bad_dates:
            issues.append({
                "type": "invalid_dates",
                "severity": "high",
                "detail": f"无法解析的日期值: {bad_dates[:5]}",
            })

    if col_type == "numeric":
        nums = []
        non_numeric = []
        for v in non_empty:
            try:
                nums.append(float(v.replace(",", "").strip()))
            except ValueError:
                non_numeric.append(v)
        if non_numeric:
            issues.append({
                "type": "non_numeric_values",
                "severity": "high",
                "detail": f"数值列中包含非数值: {non_numeric[:5]}",
            })
        if nums:
            neg = [n for n in nums if n < 0]
            if neg:
                issues.append({
                    "type": "negative_values",
                    "severity": "medium",
                    "detail": f"检测到 {len(neg)} 个负值（最小: {min(neg)}）",
                })

    if col_type == "text":
        # 大小写不一致检测
        lower_groups = {}
        for v in non_empty:
            key = v.strip().lower()
            if key not in lower_groups:
                lower_groups[key] = set()
            lower_groups[key].add(v.strip())
        inconsistent = {k: list(v) for k, v in lower_groups.items() if len(v) > 1}
        if inconsistent:
            issues.append({
                "type": "case_inconsistency",
                "severity": "low",
                "detail": f"大小写不一致: {dict(list(inconsistent.items())[:3])}",
            })

        # 前后空格检测
        whitespace_issues = [v for v in values if v != v.strip() and v.strip()]
        if whitespace_issues:
            issues.append({
                "type": "whitespace",
                "severity": "low",
                "detail": f"{len(whitespace_issues)} 个值有前后空格",
            })

    if missing > 0:
        # 区分真空值和占位符
        placeholders = [v for v in values if v.strip().lower() in PLACEHOLDER_VALUES and v.strip()]
        true_empty = [v for v in values if not v.strip()]
        detail_parts = []
        if true_empty:
            detail_parts.append(f"{len(true_empty)} 个空值")
        if placeholders:
            placeholder_counts = Counter(v.strip() for v in placeholders)
            detail_parts.append(f"占位符: {dict(placeholder_counts)}")
        issues.append({
            "type": "missing_values",
            "severity": "medium" if missing / total > 0.1 else "low",
            "detail": ", ".join(detail_parts) if detail_parts else f"{missing} 个缺失",
        })

    profile["issues"] = issues
    profile["sample_values"] = unique_vals[:10]
    return profile


def find_duplicates(rows, headers):
    """检测重复行（完全重复 + 近似重复）"""
    # 完全重复（所有字段相同）
    row_strs = []
    for r in rows:
        row_strs.append(json.dumps(r, ensure_ascii=False, sort_keys=True))

    exact_dupes = []
    seen = {}
    for i, s in enumerate(row_strs):
        if s in seen:
            exact_dupes.append({"row": i + 2, "duplicate_of": seen[s] + 2})  # +2 for header + 0-index
        else:
            seen[s] = i

    # 近似重复（排除第一列/ID列，其余字段相同）
    near_dupes = []
    if len(headers) > 1:
        non_id_cols = headers[1:]  # 假设第一列是 ID
        seen_partial = {}
        for i, r in enumerate(rows):
            key = json.dumps({k: r[k] for k in non_id_cols}, ensure_ascii=False, sort_keys=True)
            if key in seen_partial:
                near_dupes.append({
                    "row": i + 2,
                    "similar_to": seen_partial[key] + 2,
                    "matching_fields": non_id_cols,
                })
            else:
                seen_partial[key] = i

    return exact_dupes, near_dupes


def cmd_profile(filepath, output_format="json"):
    """生成数据质量报告"""
    if not os.path.exists(filepath):
        print(json.dumps({"error": f"文件不存在: {filepath}"}))
        return 1

    try:
        headers, rows = read_data(filepath)
    except ImportError as e:
        print(json.dumps({"error": str(e)}))
        return 1
    if not rows:
        print(json.dumps({"error": "文件为空或无有效数据"}))
        return 1

    # 列画像
    columns = []
    all_issues = []
    for h in headers:
        values = [r[h] for r in rows]
        col_profile = profile_column(h, values)
        columns.append(col_profile)
        for issue in col_profile["issues"]:
            issue["column"] = h
            all_issues.append(issue)

    # 重复检测
    exact_dupes, near_dupes = find_duplicates(rows, headers)
    if exact_dupes:
        all_issues.append({
            "type": "exact_duplicates",
            "severity": "high",
            "column": "*",
            "detail": f"{len(exact_dupes)} 行完全重复",
        })
    if near_dupes:
        all_issues.append({
            "type": "near_duplicates",
            "severity": "medium",
            "column": "*",
            "detail": f"{len(near_dupes)} 行近似重复（仅ID不同）",
        })

    # 质量评分
    high_count = sum(1 for i in all_issues if i["severity"] == "high")
    med_count = sum(1 for i in all_issues if i["severity"] == "medium")
    low_count = sum(1 for i in all_issues if i["severity"] == "low")
    score = max(0, 100 - high_count * 15 - med_count * 8 - low_count * 3)

    report = {
        "file": os.path.basename(filepath),
        "rows": len(rows),
        "columns": len(headers),
        "quality_score": score,
        "issues_summary": {
            "high": high_count,
            "medium": med_count,
            "low": low_count,
            "total": len(all_issues),
        },
        "issues": sorted(all_issues, key=lambda x: {"high": 0, "medium": 1, "low": 2}[x["severity"]]),
        "column_profiles": columns,
        "exact_duplicates": exact_dupes,
        "near_duplicates": near_dupes,
        "sample_rows": rows[:5],
    }

    # 保存报告
    ensure_workspace()
    report_file = os.path.join(WORKSPACE, "latest_profile.json")
    with open(report_file, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    audit_log("profile", {"file": filepath, "score": score, "issues": len(all_issues)})

    if output_format == "json":
        print(json.dumps(report, ensure_ascii=False, indent=2))
    else:
        # 简洁文本摘要（给 LLM 消费）
        print(f"文件: {report['file']} ({report['rows']}行 × {report['columns']}列)")
        print(f"质量评分: {score}/100")
        print(f"问题: {high_count} high, {med_count} medium, {low_count} low")
        for issue in all_issues:
            severity_icon = {"high": "🔴", "medium": "🟡", "low": "🟢"}[issue["severity"]]
            print(f"  {severity_icon} [{issue['column']}] {issue['type']}: {issue['detail']}")

    return 0


# ── execute 子命令 ────────────────────────────────────────

def op_dedup(headers, rows, args):
    """去重操作"""
    original_count = len(rows)
    # 完全去重
    seen = set()
    deduped = []
    for r in rows:
        key = json.dumps(r, ensure_ascii=False, sort_keys=True)
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    removed = original_count - len(deduped)
    return deduped, {
        "operation": "dedup",
        "rows_before": original_count,
        "rows_after": len(deduped),
        "rows_removed": removed,
    }


def op_dedup_near(headers, rows, args):
    """近似去重（保留第一条）"""
    if len(headers) < 2:
        return rows, {"operation": "dedup_near", "skipped": "列数不足"}

    original_count = len(rows)
    non_id_cols = headers[1:]
    seen = set()
    deduped = []
    for r in rows:
        key = json.dumps({k: r[k] for k in non_id_cols}, ensure_ascii=False, sort_keys=True)
        if key not in seen:
            seen.add(key)
            deduped.append(r)

    removed = original_count - len(deduped)
    return deduped, {
        "operation": "dedup_near",
        "rows_before": original_count,
        "rows_after": len(deduped),
        "rows_removed": removed,
    }


def op_trim(headers, rows, args):
    """去除所有文本值的前后空格"""
    changes = 0
    for r in rows:
        for h in headers:
            trimmed = r[h].strip()
            if trimmed != r[h]:
                r[h] = trimmed
                changes += 1
    return rows, {"operation": "trim", "cells_trimmed": changes}


def op_fix_dates(headers, rows, args):
    """统一日期格式为 YYYY-MM-DD"""
    target_cols = args if args else [
        col for col in headers
        if detect_column_type([r[col] for r in rows]) == "datetime"
    ]

    total_fixed = 0
    unfixable = []

    for col in target_cols:
        if col not in headers:
            continue
        for r in rows:
            val = r[col].strip()
            if not val or val.lower() in PLACEHOLDER_VALUES:
                continue

            # 已经是目标格式
            if re.match(r"^\d{4}-\d{2}-\d{2}$", val):
                try:
                    datetime.strptime(val, "%Y-%m-%d")
                    continue
                except ValueError:
                    pass

            # 尝试各种格式
            parsed = None
            for pattern, fmt, label in DATE_PATTERNS:
                if re.match(pattern, val):
                    try:
                        parsed = datetime.strptime(val, fmt)
                        break
                    except ValueError:
                        continue

            if parsed:
                r[col] = parsed.strftime("%Y-%m-%d")
                total_fixed += 1
            else:
                unfixable.append({"row": rows.index(r) + 2, "column": col, "value": val})

    return rows, {
        "operation": "fix_dates",
        "columns": target_cols,
        "dates_fixed": total_fixed,
        "unfixable": unfixable[:10],
    }


def op_fix_case(headers, rows, args):
    """统一指定列的大小写（转小写）"""
    target_cols = args if args else []
    if not target_cols:
        return rows, {"operation": "fix_case", "skipped": "未指定目标列"}

    changes = 0
    for col in target_cols:
        if col not in headers:
            continue
        for r in rows:
            lower = r[col].strip().lower()
            if lower != r[col]:
                r[col] = lower
                changes += 1

    return rows, {"operation": "fix_case", "columns": target_cols, "cells_changed": changes}


def op_fill_missing(headers, rows, args):
    """标记缺失值为统一占位符"""
    changes = 0
    marker = "[MISSING]"
    for r in rows:
        for h in headers:
            val = r[h].strip()
            if not val or val.lower() in PLACEHOLDER_VALUES:
                if r[h] != marker:
                    r[h] = marker
                    changes += 1

    return rows, {"operation": "fill_missing", "marker": marker, "cells_marked": changes}


def op_remove_test(headers, rows, args):
    """移除疑似测试数据行"""
    test_keywords = {"test", "测试", "请忽略", "debug", "tmp", "temp", "foo", "bar"}
    original_count = len(rows)
    cleaned = []
    removed_rows = []
    for i, r in enumerate(rows):
        row_text = " ".join(r.values()).lower()
        if any(kw in row_text for kw in test_keywords):
            removed_rows.append(i + 2)
        else:
            cleaned.append(r)

    return cleaned, {
        "operation": "remove_test",
        "rows_before": original_count,
        "rows_after": len(cleaned),
        "removed_rows": removed_rows,
    }


# 操作注册表
OPERATIONS = {
    "dedup": {"fn": op_dedup, "desc": "完全去重", "risk": "low"},
    "dedup_near": {"fn": op_dedup_near, "desc": "近似去重（仅ID不同）", "risk": "medium"},
    "trim": {"fn": op_trim, "desc": "去除前后空格", "risk": "low"},
    "fix_dates": {"fn": op_fix_dates, "desc": "统一日期格式为 YYYY-MM-DD", "risk": "medium"},
    "fix_case": {"fn": op_fix_case, "desc": "统一大小写", "risk": "low"},
    "fill_missing": {"fn": op_fill_missing, "desc": "标记缺失值", "risk": "low"},
    "remove_test": {"fn": op_remove_test, "desc": "移除测试数据", "risk": "medium"},
}


def cmd_execute(filepath, operations, op_args=None):
    """执行清洗操作"""
    if not os.path.exists(filepath):
        print(json.dumps({"error": f"文件不存在: {filepath}"}))
        return 1

    try:
        input_fmt = detect_format(filepath)
        headers, rows = read_data(filepath, input_fmt)
    except ImportError as e:
        print(json.dumps({"error": str(e)}))
        return 1
    if not rows:
        print(json.dumps({"error": "文件为空"}))
        return 1

    # 保存原始版本
    original_file = save_version(filepath, headers, rows, "v0_original")

    results = []
    current_rows = [dict(r) for r in rows]  # deep copy

    for op_name in operations:
        if op_name not in OPERATIONS:
            results.append({"operation": op_name, "error": f"未知操作: {op_name}"})
            continue

        op_info = OPERATIONS[op_name]
        args = (op_args or {}).get(op_name, [])
        rows_before = len(current_rows)

        # 执行操作
        current_rows, op_result = op_info["fn"](headers, current_rows, args)

        # 保存版本快照
        step = len(results) + 1
        version_file = save_version(filepath, headers, current_rows, f"v{step}_{op_name}")
        op_result["version_file"] = os.path.basename(version_file)

        results.append(op_result)

    # 保存最终结果（保持原始格式输出）
    output_basename = os.path.splitext(os.path.basename(filepath))[0]
    ext_map = {"csv": ".csv", "tsv": ".tsv", "json": ".json", "jsonl": ".jsonl", "xlsx": ".xlsx"}
    output_ext = ext_map.get(input_fmt, ".csv")
    output_file = os.path.join(WORKSPACE, f"{output_basename}_cleaned{output_ext}")
    write_data(output_file, headers, current_rows, input_fmt)

    # 生成报告
    report = {
        "input": os.path.basename(filepath),
        "output": output_file,
        "original_rows": len(rows),
        "final_rows": len(current_rows),
        "operations_applied": len(results),
        "steps": results,
        "original_backup": original_file,
    }

    # 保存 markdown 报告
    md_report = generate_report_md(filepath, rows, current_rows, results)
    report_file = os.path.join(WORKSPACE, "report.md")
    with open(report_file, "w", encoding="utf-8") as f:
        f.write(md_report)
    report["report_file"] = report_file

    audit_log("execute", {
        "file": filepath,
        "operations": operations,
        "rows_before": len(rows),
        "rows_after": len(current_rows),
    })

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


def generate_report_md(filepath, original_rows, cleaned_rows, steps):
    """生成 Markdown 清洗报告"""
    lines = [
        f"# 数据清洗报告",
        f"",
        f"**文件**: {os.path.basename(filepath)}",
        f"**时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"**行数变化**: {len(original_rows)} → {len(cleaned_rows)} "
        f"({'减少' if len(cleaned_rows) < len(original_rows) else '不变'} "
        f"{abs(len(original_rows) - len(cleaned_rows))} 行)",
        f"",
        f"## 操作步骤",
        f"",
    ]

    for i, step in enumerate(steps, 1):
        op = step.get("operation", "unknown")
        lines.append(f"### {i}. {op}")
        for k, v in step.items():
            if k in ("operation", "version_file"):
                continue
            lines.append(f"- **{k}**: {v}")
        lines.append("")

    lines.extend([
        f"## 版本历史",
        f"",
        f"所有版本快照保存在 `{VERSION_DIR}/`",
        f"如需回滚，可从对应版本文件恢复。",
    ])

    return "\n".join(lines)


# ── validate 子命令 ───────────────────────────────────────

def cmd_validate(original_path, cleaned_path):
    """清洗前后对比验证"""
    if not os.path.exists(original_path) or not os.path.exists(cleaned_path):
        print(json.dumps({"error": "文件不存在"}))
        return 1

    try:
        orig_headers, orig_rows = read_data(original_path)
        clean_headers, clean_rows = read_data(cleaned_path)
    except ImportError as e:
        print(json.dumps({"error": str(e)}))
        return 1

    report = {
        "original": {"rows": len(orig_rows), "hash": file_hash(original_path)},
        "cleaned": {"rows": len(clean_rows), "hash": file_hash(cleaned_path)},
        "row_diff": len(clean_rows) - len(orig_rows),
        "checks": [],
    }

    # 检查: 行数是否合理（不应删除超过 50%）
    if len(clean_rows) < len(orig_rows) * 0.5:
        report["checks"].append({
            "check": "row_count",
            "status": "WARNING",
            "detail": f"行数减少超过50%: {len(orig_rows)} → {len(clean_rows)}",
        })
    else:
        report["checks"].append({
            "check": "row_count",
            "status": "OK",
            "detail": f"{len(orig_rows)} → {len(clean_rows)}",
        })

    # 检查: 列数应不变
    clean_headers_list = clean_headers
    if set(orig_headers) != set(clean_headers_list):
        report["checks"].append({
            "check": "column_integrity",
            "status": "ERROR",
            "detail": "列结构发生变化",
        })
    else:
        report["checks"].append({
            "check": "column_integrity",
            "status": "OK",
            "detail": "列结构完整",
        })

    all_ok = all(c["status"] == "OK" for c in report["checks"])
    report["verdict"] = "PASS" if all_ok else "REVIEW_NEEDED"

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


# ── history 子命令 ────────────────────────────────────────

def cmd_history(filepath):
    """查看清洗版本历史"""
    ensure_workspace()
    basename = os.path.splitext(os.path.basename(filepath))[0]

    versions = []
    if os.path.exists(VERSION_DIR):
        for f in sorted(os.listdir(VERSION_DIR)):
            if f.startswith(basename) and f.endswith(".csv"):
                full_path = os.path.join(VERSION_DIR, f)
                _, rows = read_csv(full_path)
                versions.append({
                    "file": f,
                    "rows": len(rows),
                    "hash": file_hash(full_path),
                    "size": os.path.getsize(full_path),
                })

    # 审计日志
    logs = []
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, "r") as fh:
            for line in fh:
                try:
                    entry = json.loads(line.strip())
                    if basename in entry.get("file", ""):
                        logs.append(entry)
                except json.JSONDecodeError:
                    continue

    print(json.dumps({
        "file": os.path.basename(filepath),
        "versions": versions,
        "audit_log": logs[-10:],
    }, ensure_ascii=False, indent=2))
    return 0


# ── list-ops 子命令 ──────────────────────────────────────

def cmd_list_ops():
    """列出所有可用的清洗操作"""
    ops = []
    for name, info in OPERATIONS.items():
        ops.append({
            "name": name,
            "description": info["desc"],
            "risk": info["risk"],
        })
    print(json.dumps({"operations": ops}, ensure_ascii=False, indent=2))
    return 0


# ── 主入口 ───────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="数据清洗 CLI 工具（支持 CSV/TSV/JSON/JSONL/Excel）"
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # profile
    p_profile = subparsers.add_parser("profile", help="数据画像（质量报告）")
    p_profile.add_argument("file", help="数据文件路径（CSV/TSV/JSON/JSONL/XLSX）")
    p_profile.add_argument("--format", choices=["json", "text"], default="json")

    # execute
    p_exec = subparsers.add_parser("execute", help="执行清洗操作")
    p_exec.add_argument("file", help="数据文件路径（CSV/TSV/JSON/JSONL/XLSX）")
    p_exec.add_argument("--ops", nargs="+", required=True,
                        help=f"操作列表: {', '.join(OPERATIONS.keys())}")
    p_exec.add_argument("--fix-case-cols", nargs="+", default=[],
                        help="fix_case 操作的目标列")
    p_exec.add_argument("--fix-date-cols", nargs="+", default=[],
                        help="fix_dates 操作的目标列（默认自动检测）")

    # validate
    p_validate = subparsers.add_parser("validate", help="清洗前后对比验证")
    p_validate.add_argument("original", help="原始文件")
    p_validate.add_argument("cleaned", help="清洗后文件")

    # history
    p_history = subparsers.add_parser("history", help="查看版本历史")
    p_history.add_argument("file", help="原始文件路径")

    # list-ops
    subparsers.add_parser("list-ops", help="列出所有可用操作")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return 1

    if args.command == "profile":
        return cmd_profile(args.file, args.format)
    elif args.command == "execute":
        op_args = {}
        if args.fix_case_cols:
            op_args["fix_case"] = args.fix_case_cols
        if args.fix_date_cols:
            op_args["fix_dates"] = args.fix_date_cols
        return cmd_execute(args.file, args.ops, op_args)
    elif args.command == "validate":
        return cmd_validate(args.original, args.cleaned)
    elif args.command == "history":
        return cmd_history(args.file)
    elif args.command == "list-ops":
        return cmd_list_ops()

    return 0


if __name__ == "__main__":
    sys.exit(main())
